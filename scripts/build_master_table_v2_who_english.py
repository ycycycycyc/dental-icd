#!/usr/bin/env python3
"""Add WHO ICD-10 English terminology mappings to the K00-K14 master table."""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


INPUT_COLUMNS = [
    "chapter",
    "section",
    "category_code",
    "category_name_cn",
    "subcategory_code",
    "subcategory_name_cn",
    "diagnosis_code",
    "diagnosis_name_cn",
    "chapter_code",
    "parent_code",
    "is_subtype",
    "subtype_number",
    "code_level",
]

ADDED_COLUMNS = [
    "diagnosis_name_en",
    "category_name_en",
    "subcategory_name_en",
    "english_mapping_confidence",
]

OUTPUT_COLUMNS = INPUT_COLUMNS + ADDED_COLUMNS

DEFAULT_INPUT = Path("K00-K14_master_table_v1_english_columns.tsv")
DEFAULT_TSV = Path("K00-K14_master_table_v2.tsv")
DEFAULT_XLSX = Path("K00-K14_master_table_v2.xlsx")

K00_K14_RE = re.compile(r"^K(0[0-9]|1[0-4])")

# Source: WHO ICD-10 Version:2019 browser, public K00-K14 block.
# https://icd.who.int/browse10/2019/en#/K00-K14
# Parsed from:
# https://icd.who.int/browse10/2019/en/GetConcept?ConceptId=K00-K14
WHO_TERMS = {
    "K00": "Disorders of tooth development and eruption",
    "K00.0": "Anodontia",
    "K00.1": "Supernumerary teeth",
    "K00.2": "Abnormalities of size and form of teeth",
    "K00.3": "Mottled teeth",
    "K00.4": "Disturbances in tooth formation",
    "K00.5": "Hereditary disturbances in tooth structure, not elsewhere classified",
    "K00.6": "Disturbances in tooth eruption",
    "K00.7": "Teething syndrome",
    "K00.8": "Other disorders of tooth development",
    "K00.9": "Disorder of tooth development, unspecified",
    "K01": "Embedded and impacted teeth",
    "K01.0": "Embedded teeth",
    "K01.1": "Impacted teeth",
    "K02": "Dental caries",
    "K02.0": "Caries limited to enamel",
    "K02.1": "Caries of dentine",
    "K02.2": "Caries of cementum",
    "K02.3": "Arrested dental caries",
    "K02.4": "Odontoclasia",
    "K02.5": "Caries with pulp exposure",
    "K02.8": "Other dental caries",
    "K02.9": "Dental caries, unspecified",
    "K03": "Other diseases of hard tissues of teeth",
    "K03.0": "Excessive attrition of teeth",
    "K03.1": "Abrasion of teeth",
    "K03.2": "Erosion of teeth",
    "K03.3": "Pathological resorption of teeth",
    "K03.4": "Hypercementosis",
    "K03.5": "Ankylosis of teeth",
    "K03.6": "Deposits [accretions] on teeth",
    "K03.7": "Posteruptive colour changes of dental hard tissues",
    "K03.8": "Other specified diseases of hard tissues of teeth",
    "K03.9": "Disease of hard tissues of teeth, unspecified",
    "K04": "Diseases of pulp and periapical tissues",
    "K04.0": "Pulpitis",
    "K04.1": "Necrosis of pulp",
    "K04.2": "Pulp degeneration",
    "K04.3": "Abnormal hard tissue formation in pulp",
    "K04.4": "Acute apical periodontitis of pulpal origin",
    "K04.5": "Chronic apical periodontitis",
    "K04.6": "Periapical abscess with sinus",
    "K04.7": "Periapical abscess without sinus",
    "K04.8": "Radicular cyst",
    "K04.9": "Other and unspecified diseases of pulp and periapical tissues",
    "K05": "Gingivitis and periodontal diseases",
    "K05.0": "Acute gingivitis",
    "K05.1": "Chronic gingivitis",
    "K05.2": "Acute periodontitis",
    "K05.3": "Chronic periodontitis",
    "K05.4": "Periodontosis",
    "K05.5": "Other periodontal diseases",
    "K05.6": "Periodontal disease, unspecified",
    "K06": "Other disorders of gingiva and edentulous alveolar ridge",
    "K06.0": "Gingival recession",
    "K06.1": "Gingival enlargement",
    "K06.2": "Gingival and edentulous alveolar ridge lesions associated with trauma",
    "K06.8": "Other specified disorders of gingiva and edentulous alveolar ridge",
    "K06.9": "Disorder of gingiva and edentulous alveolar ridge, unspecified",
    "K07": "Dentofacial anomalies [including malocclusion]",
    "K07.0": "Major anomalies of jaw size",
    "K07.1": "Anomalies of jaw-cranial base relationship",
    "K07.2": "Anomalies of dental arch relationship",
    "K07.3": "Anomalies of tooth position",
    "K07.4": "Malocclusion, unspecified",
    "K07.5": "Dentofacial functional abnormalities",
    "K07.6": "Temporomandibular joint disorders",
    "K07.8": "Other dentofacial anomalies",
    "K07.9": "Dentofacial anomaly, unspecified",
    "K08": "Other disorders of teeth and supporting structures",
    "K08.0": "Exfoliation of teeth due to systemic causes",
    "K08.1": "Loss of teeth due to accident, extraction or local periodontal disease",
    "K08.2": "Atrophy of edentulous alveolar ridge",
    "K08.3": "Retained dental root",
    "K08.8": "Other specified disorders of teeth and supporting structures",
    "K08.9": "Disorder of teeth and supporting structures, unspecified",
    "K09": "Cysts of oral region, not elsewhere classified",
    "K09.0": "Developmental odontogenic cysts",
    "K09.1": "Developmental (nonodontogenic) cysts of oral region",
    "K09.2": "Other cysts of jaw",
    "K09.8": "Other cysts of oral region, not elsewhere classified",
    "K09.9": "Cyst of oral region, unspecified",
    "K10": "Other diseases of jaws",
    "K10.0": "Developmental disorders of jaws",
    "K10.1": "Giant cell granuloma, central",
    "K10.2": "Inflammatory conditions of jaws",
    "K10.3": "Alveolitis of jaws",
    "K10.8": "Other specified diseases of jaws",
    "K10.9": "Disease of jaws, unspecified",
    "K11": "Diseases of salivary glands",
    "K11.0": "Atrophy of salivary gland",
    "K11.1": "Hypertrophy of salivary gland",
    "K11.2": "Sialoadenitis",
    "K11.3": "Abscess of salivary gland",
    "K11.4": "Fistula of salivary gland",
    "K11.5": "Sialolithiasis",
    "K11.6": "Mucocele of salivary gland",
    "K11.7": "Disturbances of salivary secretion",
    "K11.8": "Other diseases of salivary glands",
    "K11.9": "Disease of salivary gland, unspecified",
    "K12": "Stomatitis and related lesions",
    "K12.0": "Recurrent oral aphthae",
    "K12.1": "Other forms of stomatitis",
    "K12.2": "Cellulitis and abscess of mouth",
    "K12.3": "Oral mucositis (ulcerative)",
    "K13": "Other diseases of lip and oral mucosa",
    "K13.0": "Diseases of lips",
    "K13.1": "Cheek and lip biting",
    "K13.2": "Leukoplakia and other disturbances of oral epithelium, including tongue",
    "K13.3": "Hairy leukoplakia",
    "K13.4": "Granuloma and granuloma-like lesions of oral mucosa",
    "K13.5": "Oral submucous fibrosis",
    "K13.6": "Irritative hyperplasia of oral mucosa",
    "K13.7": "Other and unspecified lesions of oral mucosa",
    "K14": "Diseases of tongue",
    "K14.0": "Glossitis",
    "K14.1": "Geographic tongue",
    "K14.2": "Median rhomboid glossitis",
    "K14.3": "Hypertrophy of tongue papillae",
    "K14.4": "Atrophy of tongue papillae",
    "K14.5": "Plicated tongue",
    "K14.6": "Glossodynia",
    "K14.8": "Other diseases of tongue",
    "K14.9": "Disease of tongue, unspecified",
}


def is_canonical_subcategory_diagnosis(row: dict[str, str]) -> bool:
    return row["diagnosis_code"] == f"{row['subcategory_code']}00"


def read_input(input_path: Path) -> list[dict[str, str]]:
    with input_path.open(encoding="utf-8", newline="") as input_file:
        reader = csv.DictReader(input_file, delimiter="\t")
        if reader.fieldnames != INPUT_COLUMNS:
            raise ValueError(
                f"Unexpected input columns: {reader.fieldnames}. "
                f"Expected: {INPUT_COLUMNS}"
            )
        return [dict(row) for row in reader]


def add_english_mappings(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    output_rows: list[dict[str, str]] = []
    by_code: dict[str, dict[str, str]] = {}

    for source in rows:
        row = dict(source)
        category_name_en = WHO_TERMS[row["category_code"]]
        subcategory_name_en = WHO_TERMS[row["subcategory_code"]]

        row["category_name_en"] = category_name_en
        row["subcategory_name_en"] = subcategory_name_en

        if row["is_subtype"] == "TRUE":
            row["diagnosis_name_en"] = ""
            row["english_mapping_confidence"] = "LOW"
        elif is_canonical_subcategory_diagnosis(row):
            row["diagnosis_name_en"] = subcategory_name_en
            row["english_mapping_confidence"] = "HIGH"
        elif row["diagnosis_code"] in WHO_TERMS:
            row["diagnosis_name_en"] = WHO_TERMS[row["diagnosis_code"]]
            row["english_mapping_confidence"] = "HIGH"
        else:
            row["diagnosis_name_en"] = subcategory_name_en
            row["english_mapping_confidence"] = "MEDIUM"

        output_rows.append(row)
        by_code[row["diagnosis_code"]] = row

    for row in output_rows:
        if row["is_subtype"] == "TRUE":
            parent = by_code.get(row["parent_code"])
            row["diagnosis_name_en"] = (
                parent["diagnosis_name_en"] if parent else row["subcategory_name_en"]
            )

    return output_rows


def write_tsv(rows: list[dict[str, str]], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.DictWriter(
            output_file,
            fieldnames=OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "K00-K14 Master v2"
    worksheet.append(OUTPUT_COLUMNS)
    for row in rows:
        worksheet.append([row[column] for column in OUTPUT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 28,
        "B": 28,
        "C": 16,
        "D": 24,
        "E": 18,
        "F": 34,
        "G": 18,
        "H": 38,
        "I": 14,
        "J": 14,
        "K": 12,
        "L": 15,
        "M": 14,
        "N": 44,
        "O": 38,
        "P": 44,
        "Q": 24,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    table = Table(displayName="K00K14MasterV2", ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    workbook.save(output_path)


def validate(input_rows: list[dict[str, str]], output_rows: list[dict[str, str]]) -> None:
    input_subtypes = {
        row["diagnosis_code"] for row in input_rows if row["is_subtype"] == "TRUE"
    }
    output_subtypes = {
        row["diagnosis_code"] for row in output_rows if row["is_subtype"] == "TRUE"
    }
    codes = [row["diagnosis_code"] for row in output_rows]
    missing_codes = [row for row in output_rows if not row["diagnosis_code"]]
    outside_range = [code for code in codes if not K00_K14_RE.match(code)]

    print(f"Input row count: {len(input_rows)}")
    print(f"Output row count: {len(output_rows)}")
    print(f"Missing diagnosis_code: {len(missing_codes)}")
    print(f"Rows outside K00-K14: {len(outside_range)}")
    print(f"Dropped subtype rows: {len(input_subtypes - output_subtypes)}")
    print("Counts by english_mapping_confidence:")
    for key, count in sorted(
        Counter(row["english_mapping_confidence"] for row in output_rows).items()
    ):
        print(f"{key}\t{count}")


def validate_xlsx(output_path: Path) -> None:
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    worksheet = workbook["K00-K14 Master v2"]
    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    print(f"XLSX columns match: {header == OUTPUT_COLUMNS}")
    print(f"XLSX row count: {len(rows)}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build K00-K14 master table v2 with WHO ICD-10 English mappings."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--tsv", type=Path, default=DEFAULT_TSV)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()

    input_rows = read_input(args.input)
    output_rows = add_english_mappings(input_rows)
    write_tsv(output_rows, args.tsv)
    write_xlsx(output_rows, args.xlsx)
    print(f"Wrote {args.tsv}")
    print(f"Wrote {args.xlsx}")
    validate(input_rows, output_rows)
    validate_xlsx(args.xlsx)


if __name__ == "__main__":
    main()
