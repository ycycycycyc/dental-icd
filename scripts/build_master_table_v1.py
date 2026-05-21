#!/usr/bin/env python3
"""Build the cleaned K00-K14 Chinese ICD hierarchy master table v1."""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


INPUT_COLUMNS = [
    "章",
    "节",
    "类目代码",
    "类目名称",
    "亚目代码",
    "亚目名称",
    "诊断代码",
    "诊断名称",
]

DERIVED_COLUMNS = [
    "chapter_code",
    "parent_code",
    "is_subtype",
    "subtype_number",
    "code_level",
]

OUTPUT_COLUMNS = INPUT_COLUMNS + DERIVED_COLUMNS

DEFAULT_INPUT = Path("k00_k14_full_hierarchy.tsv")
DEFAULT_TSV = Path("K00-K14_master_table_v1.tsv")
DEFAULT_XLSX = Path("K00-K14_master_table_v1.xlsx")

K00_K14_RE = re.compile(r"^K(0[0-9]|1[0-4])")
SUBTYPE_RE = re.compile(r"x(\d+)$")


def code_level(row: dict[str, str], is_subtype: bool) -> str:
    if is_subtype:
        return "subtype"
    if row["诊断代码"] == row["类目代码"]:
        return "category"
    if row["诊断代码"] == row["亚目代码"]:
        return "subcategory"
    return "diagnosis"


def build_rows(input_path: Path) -> list[dict[str, str]]:
    with input_path.open(encoding="utf-8", newline="") as input_file:
        reader = csv.DictReader(input_file, delimiter="\t")
        if reader.fieldnames != INPUT_COLUMNS:
            raise ValueError(
                f"Unexpected input columns: {reader.fieldnames}. "
                f"Expected: {INPUT_COLUMNS}"
            )

        rows: list[dict[str, str]] = []
        for source in reader:
            row = {column: source[column] for column in INPUT_COLUMNS}
            diagnosis_code = row["诊断代码"]
            subtype_match = SUBTYPE_RE.search(diagnosis_code)
            is_subtype = subtype_match is not None

            row["chapter_code"] = diagnosis_code[:3]
            row["parent_code"] = (
                diagnosis_code[: subtype_match.start()]
                if subtype_match
                else row["亚目代码"]
            )
            row["is_subtype"] = "TRUE" if is_subtype else "FALSE"
            row["subtype_number"] = subtype_match.group(1) if subtype_match else ""
            row["code_level"] = code_level(row, is_subtype)
            rows.append(row)

    return rows


def write_tsv(rows: list[dict[str, str]], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.DictWriter(
            output_file,
            fieldnames=OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "K00-K14 Master"
    worksheet.append(OUTPUT_COLUMNS)
    for row in rows:
        worksheet.append([row[column] for column in OUTPUT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 28,
        "B": 28,
        "C": 12,
        "D": 24,
        "E": 12,
        "F": 34,
        "G": 18,
        "H": 38,
        "I": 14,
        "J": 14,
        "K": 12,
        "L": 15,
        "M": 14,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    table = Table(displayName="K00K14MasterV1", ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    workbook.save(output_path)


def validate_rows(rows: list[dict[str, str]]) -> None:
    codes = [row["诊断代码"] for row in rows]
    code_counts = Counter(codes)
    duplicate_codes = [code for code, count in code_counts.items() if count > 1]
    outside_range = [code for code in codes if not K00_K14_RE.match(code)]
    missing_codes = [row for row in rows if not row["诊断代码"]]
    available_codes = set(codes)
    subtype_parent_errors = [
        (row["诊断代码"], row["parent_code"])
        for row in rows
        if row["is_subtype"] == "TRUE" and row["parent_code"] not in available_codes
    ]

    print(f"Total row count: {len(rows)}")
    print(f"Missing 诊断代码: {len(missing_codes)}")
    print(f"Rows outside K00-K14: {len(outside_range)}")
    print(f"Duplicate 诊断代码: {len(duplicate_codes)}")
    print(f"Subtype rows with missing parent_code: {len(subtype_parent_errors)}")
    print("Counts by chapter_code:")
    for key, count in sorted(Counter(row["chapter_code"] for row in rows).items()):
        print(f"{key}\t{count}")
    print("Counts by is_subtype:")
    for key, count in sorted(Counter(row["is_subtype"] for row in rows).items()):
        print(f"{key}\t{count}")


def validate_xlsx(output_path: Path) -> None:
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    worksheet = workbook["K00-K14 Master"]
    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    print(f"XLSX columns match: {header == OUTPUT_COLUMNS}")
    print(f"XLSX row count: {len(rows)}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build K00-K14_master_table_v1 from the official hierarchy TSV."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--tsv", type=Path, default=DEFAULT_TSV)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()

    rows = build_rows(args.input)
    write_tsv(rows, args.tsv)
    write_xlsx(rows, args.xlsx)
    print(f"Wrote {args.tsv}")
    print(f"Wrote {args.xlsx}")
    validate_rows(rows)
    validate_xlsx(args.xlsx)


if __name__ == "__main__":
    main()
