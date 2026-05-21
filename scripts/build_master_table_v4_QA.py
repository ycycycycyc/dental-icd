#!/usr/bin/env python3
"""Append QA review columns to the frozen K00-K14 v3 semantic master table."""

from __future__ import annotations

import argparse
import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


INPUT_COLUMNS = [
    "chapter",
    "section",
    "category_code",
    "category_name_cn",
    "subcategory_code",
    "subcategory_name_cn",
    "diagnosis_code",
    "diagnosis_name_cn",
    "chapter_code",
    "parent_code",
    "is_subtype",
    "subtype_number",
    "code_level",
    "diagnosis_name_en",
    "category_name_en",
    "subcategory_name_en",
    "english_mapping_confidence",
    "who_icd_code",
    "who_name_en",
    "structural_name_en",
    "semantic_name_en",
    "english_mapping_type",
    "semantic_source",
]

QA_COLUMNS = ["qa_status", "qa_reason", "qa_priority"]
OUTPUT_COLUMNS = INPUT_COLUMNS + QA_COLUMNS

DEFAULT_INPUT = Path("K00-K14_master_table_v3_semantic_en.tsv")
DEFAULT_TSV = Path("K00-K14_master_table_v4_QA.tsv")
DEFAULT_XLSX = Path("K00-K14_master_table_v4_QA.xlsx")

GENERIC_SEMANTIC_TERMS = {
    "Disease of tongue, unspecified",
    "Other and unspecified lesions of oral mucosa",
    "Other diseases of tongue",
    "Other specified diseases of hard tissues of teeth",
    "Other specified disorders of gingiva and edentulous alveolar ridge",
    "Other specified diseases of jaws",
    "Other diseases of salivary glands",
    "Other forms of stomatitis",
    "Other dental caries",
    "Disease of jaws, unspecified",
    "Disease of salivary gland, unspecified",
    "Dental caries, unspecified",
    "Dentofacial anomaly, unspecified",
    "Disorder of gingiva and edentulous alveolar ridge, unspecified",
    "Disorder of teeth and supporting structures, unspecified",
    "Disorder of tooth development, unspecified",
}


def read_input(input_path: Path) -> list[dict[str, str]]:
    with input_path.open(encoding="utf-8", newline="") as input_file:
        reader = csv.DictReader(input_file, delimiter="\t")
        if reader.fieldnames != INPUT_COLUMNS:
            raise ValueError(
                f"Unexpected input columns: {reader.fieldnames}. "
                f"Expected: {INPUT_COLUMNS}"
            )
        return [dict(row) for row in reader]


def priority_rank(priority: str) -> int:
    return {"LOW": 1, "MEDIUM": 2, "HIGH": 3}[priority]


def add_reason(
    reasons: list[str],
    priorities: list[str],
    reason: str,
    priority: str,
) -> None:
    reasons.append(reason)
    priorities.append(priority)


def semantic_too_generic(row: dict[str, str]) -> bool:
    if row["semantic_name_en"] in GENERIC_SEMANTIC_TERMS:
        return True
    if row["semantic_name_en"].startswith(("Other ", "Unspecified ")):
        return True
    if row["semantic_name_en"].endswith(", unspecified"):
        return True
    return False


def qa_row(row: dict[str, str], diagnosis_codes: set[str]) -> dict[str, str]:
    output = dict(row)
    reasons: list[str] = []
    priorities: list[str] = []

    if row["english_mapping_confidence"] == "LOW":
        add_reason(reasons, priorities, "LOW confidence English mapping", "MEDIUM")
    if row["english_mapping_type"] == "CLINICAL_TRANSLATION":
        add_reason(reasons, priorities, "Clinical translation requires review", "MEDIUM")
    if row["semantic_source"] == "CLINICAL_TRANSLATION":
        add_reason(reasons, priorities, "Semantic source is clinical translation", "MEDIUM")
    if (
        row["is_subtype"] == "TRUE"
        and row["semantic_name_en"] == row["structural_name_en"]
    ):
        add_reason(
            reasons,
            priorities,
            "Subtype semantic name matches structural parent term",
            "MEDIUM",
        )
    if semantic_too_generic(row):
        add_reason(
            reasons,
            priorities,
            "Semantic English term appears generic for Chinese diagnosis",
            "LOW",
        )
    if not row["who_icd_code"]:
        add_reason(reasons, priorities, "Missing who_icd_code", "HIGH")
    if not row["parent_code"]:
        add_reason(reasons, priorities, "Missing parent_code", "HIGH")
    if row["is_subtype"] == "TRUE" and row["parent_code"] not in diagnosis_codes:
        add_reason(reasons, priorities, "Subtype parent_code does not exist", "HIGH")

    output["qa_status"] = "NEEDS_REVIEW" if reasons else "PASS"
    output["qa_reason"] = "; ".join(reasons)
    output["qa_priority"] = (
        max(priorities, key=priority_rank) if priorities else "LOW"
    )
    return output


def add_qa_layer(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    diagnosis_codes = {row["diagnosis_code"] for row in rows}
    return [qa_row(row, diagnosis_codes) for row in rows]


def write_tsv(rows: list[dict[str, str]], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.DictWriter(
            output_file,
            fieldnames=OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "K00-K14 Master v4 QA"
    worksheet.append(OUTPUT_COLUMNS)
    for row in rows:
        worksheet.append([row[column] for column in OUTPUT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    pass_fill = PatternFill("solid", fgColor="E2F0D9")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    qa_status_col = OUTPUT_COLUMNS.index("qa_status") + 1
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[qa_status_col - 1].fill = (
            review_fill if row[qa_status_col - 1].value == "NEEDS_REVIEW" else pass_fill
        )

    for index, column in enumerate(OUTPUT_COLUMNS, start=1):
        width = 18
        if column.endswith("_name_cn") or column.endswith("_name_en"):
            width = 38
        if column in {"chapter", "section", "structural_name_en", "semantic_name_en"}:
            width = 42
        if column == "qa_reason":
            width = 58
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    table = Table(displayName="K00K14MasterV4QA", ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    workbook.save(output_path)


def validate(input_rows: list[dict[str, str]], output_rows: list[dict[str, str]]) -> None:
    input_codes = [row["diagnosis_code"] for row in input_rows]
    output_codes = [row["diagnosis_code"] for row in output_rows]
    duplicate_codes = [
        code for code, count in Counter(output_codes).items() if count > 1
    ]

    print(f"Input row count: {len(input_rows)}")
    print(f"Output row count: {len(output_rows)}")
    print(f"Same diagnosis_code sequence: {input_codes == output_codes}")
    print(f"Missing diagnosis_code: {sum(not row['diagnosis_code'] for row in output_rows)}")
    print(f"Duplicate diagnosis_code: {len(duplicate_codes)}")
    print(
        "Existing columns unchanged: "
        f"{all(all(a[col] == b[col] for col in INPUT_COLUMNS) for a, b in zip(input_rows, output_rows))}"
    )

    print("Counts by qa_status:")
    for key, count in sorted(Counter(row["qa_status"] for row in output_rows).items()):
        print(f"{key}\t{count}")
    print("Counts by qa_priority:")
    for key, count in sorted(Counter(row["qa_priority"] for row in output_rows).items()):
        print(f"{key}\t{count}")
    print("Counts by english_mapping_type:")
    for key, count in sorted(Counter(row["english_mapping_type"] for row in output_rows).items()):
        print(f"{key}\t{count}")
    print("Counts by english_mapping_confidence:")
    for key, count in sorted(
        Counter(row["english_mapping_confidence"] for row in output_rows).items()
    ):
        print(f"{key}\t{count}")


def validate_xlsx(output_path: Path) -> None:
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    worksheet = workbook["K00-K14 Master v4 QA"]
    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    print(f"XLSX columns match: {header == OUTPUT_COLUMNS}")
    print(f"XLSX row count: {len(rows)}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build K00-K14 v4 QA review layer from frozen v3 semantic table."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--tsv", type=Path, default=DEFAULT_TSV)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()

    input_rows = read_input(args.input)
    output_rows = add_qa_layer(input_rows)
    write_tsv(output_rows, args.tsv)
    write_xlsx(output_rows, args.xlsx)
    print(f"Wrote {args.tsv}")
    print(f"Wrote {args.xlsx}")
    validate(input_rows, output_rows)
    validate_xlsx(args.xlsx)


if __name__ == "__main__":
    main()
