#!/usr/bin/env python3
"""Add a semantic English refinement layer to the K00-K14 master table."""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter
from pathlib import Path

from build_master_table_v2_who_english import WHO_TERMS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


INPUT_COLUMNS = [
    "chapter",
    "section",
    "category_code",
    "category_name_cn",
    "subcategory_code",
    "subcategory_name_cn",
    "diagnosis_code",
    "diagnosis_name_cn",
    "chapter_code",
    "parent_code",
    "is_subtype",
    "subtype_number",
    "code_level",
    "diagnosis_name_en",
    "category_name_en",
    "subcategory_name_en",
    "english_mapping_confidence",
]

ADDED_COLUMNS = [
    "who_icd_code",
    "who_name_en",
    "structural_name_en",
    "semantic_name_en",
    "english_mapping_type",
    "semantic_source",
]

OUTPUT_COLUMNS = INPUT_COLUMNS + ADDED_COLUMNS

DEFAULT_INPUT = Path("K00-K14_master_table_v2.tsv")
DEFAULT_TSV = Path("K00-K14_master_table_v3_semantic_en.tsv")
DEFAULT_XLSX = Path("K00-K14_master_table_v3_semantic_en.xlsx")

K00_K14_RE = re.compile(r"^K(0[0-9]|1[0-4])")

# Exact clinical terms for China extension rows where a recognized dental or
# oral medicine term is preferable to structural parent inheritance.
SEMANTIC_OVERRIDES = {
    "少牙畸形": "Oligodontia",
    "先天缺牙": "Congenital absence of teeth",
    "牙齿发育不全": "Dental hypoplasia",
    "牙齿缺少": "Hypodontia",
    "多生牙": "Supernumerary teeth",
    "第四臼齿": "Fourth molar",
    "巨牙症": "Macrodontia",
    "釉珠": "Enamel pearl",
    "过小牙": "Microdontia",
    "牛牙症": "Taurodontism",
    "畸形中央尖": "Dens evaginatus",
    "牙内陷": "Dens invaginatus",
    "双生牙": "Geminated tooth",
    "结合齿": "Concrescence of teeth",
    "套叠齿": "Dens invaginatus",
    "融合齿": "Fusion of teeth",
    "圆锥齿": "Conical tooth",
    "齿前突": "Dens evaginatus",
    "齿中突": "Dens invaginatus",
    "无氟釉质不透明": "Non-fluoride enamel opacity",
    "氟牙症": "Dental fluorosis",
    "弯曲牙": "Dilacerated tooth",
    "特奈牙": "Turner tooth",
    "牙根发育不良": "Root dysplasia",
    "区域性牙齿发育异常": "Regional odontodysplasia",
    "釉质发育不全（新生儿）（生后）（生前）": "Enamel hypoplasia",
    "壳状牙": "Shell teeth",
    "牙本质发育不全": "Dentinogenesis imperfecta",
    "牙生长不全": "Odontogenesis imperfecta",
    "釉质发生不全": "Amelogenesis imperfecta",
    "诞生牙": "Natal tooth",
    "恒牙萌出过迟": "Delayed eruption of permanent teeth",
    "恒牙早萌": "Premature eruption of permanent teeth",
    "低位乳牙": "Infraoccluded primary tooth",
    "个别乳磨牙早失": "Premature loss of primary molar",
    "新生儿牙": "Neonatal tooth",
    "牙齿萌出过早": "Premature tooth eruption",
    "牙齿萌出过晚": "Delayed tooth eruption",
    "乳齿过早脱落": "Premature loss of deciduous teeth",
    "乳牙滞留": "Retained deciduous tooth",
    "四环素牙": "Tetracycline-stained teeth",
    "牙齿形成期间颜色改变": "Intrinsic tooth discoloration during tooth formation",
    "牙齿白斑点损害": "White spot lesion of tooth",
    "乳牙中龋": "Moderate caries of primary teeth",
    "牙折断": "Fractured tooth",
    "婴儿黑牙病": "Infantile melanodontia",
    "黑牙折断": "Fracture of black tooth",
    "继发龋": "Secondary caries",
    "急性龋": "Acute dental caries",
    "乳牙浅龋": "Superficial caries of primary teeth",
    "乳牙深龋": "Deep caries of primary teeth",
    "恒牙浅龋": "Superficial caries of permanent teeth",
    "恒牙中龋": "Moderate caries of permanent teeth",
    "恒牙深龋": "Deep caries of permanent teeth",
    "龋病": "Dental caries",
    "蔓延性龋": "Rampant caries",
    "牙齿颌面磨损": "Occlusal tooth wear",
    "邻面磨损": "Interproximal tooth wear",
    "牙齿磨损": "Tooth wear",
    "牙齿楔状缺损": "Wedge-shaped defect of tooth",
    "净齿剂牙磨损": "Dentifrice abrasion",
    "习惯性牙磨损": "Habitual tooth abrasion",
    "职业性牙磨损": "Occupational tooth abrasion",
    "宗教仪式性牙磨损": "Ritual tooth abrasion",
    "传统性牙磨损": "Traditional tooth abrasion",
    "牙酸蚀病": "Dental erosion",
    "特发性牙腐蚀": "Idiopathic dental erosion",
    "药物性牙腐蚀": "Drug-induced dental erosion",
    "职业性牙腐蚀": "Occupational dental erosion",
    "持续性呕吐致牙腐蚀": "Dental erosion due to persistent vomiting",
    "牙髓内部肉芽肿": "Internal granuloma of pulp",
    "牙根外吸收": "External root resorption",
    "牙内吸收": "Internal resorption of tooth",
    "牙齿变色": "Tooth discoloration",
    "牙石": "Dental calculus",
    "龈下牙石（龈下垢）": "Subgingival calculus",
    "龈上牙石（龈上垢）": "Supragingival calculus",
    "牙本质过敏症": "Dentin hypersensitivity",
    "辐照性牙釉质": "Radiation-induced enamel defect",
    "牙隐裂": "Cracked tooth",
    "牙震荡": "Tooth concussion",
    "牙根纵裂": "Vertical root fracture",
    "溃疡性牙髓炎": "Ulcerative pulpitis",
    "增生性牙髓炎": "Hyperplastic pulpitis",
    "急性牙髓炎": "Acute pulpitis",
    "慢性牙髓炎": "Chronic pulpitis",
    "可逆性牙髓炎": "Reversible pulpitis",
    "不可逆性牙髓炎": "Irreversible pulpitis",
    "牙髓坏疽": "Gangrene of pulp",
    "牙髓钙化": "Pulp calcification",
    "牙髓石": "Pulp stone",
    "急性根尖周炎": "Acute apical periodontitis",
    "慢性根尖周炎": "Chronic apical periodontitis",
    "根尖肉芽肿": "Periapical granuloma",
    "根尖脓肿": "Apical abscess",
    "牙槽脓肿": "Dentoalveolar abscess",
    "剩余牙根脓肿": "Abscess of residual tooth root",
    "根尖囊肿": "Periapical cyst",
    "根尖周囊肿": "Periapical cyst",
    "残余牙根囊肿": "Residual radicular cyst",
    "牙周牙髓综合征": "Periodontal-endodontic lesion",
    "急性龈乳头炎": "Acute gingival papillitis",
    "龈炎": "Gingivitis",
    "青春期龈炎": "Puberty gingivitis",
    "龈乳头炎": "Gingival papillitis",
    "菌斑性龈炎": "Plaque-induced gingivitis",
    "萌出性龈炎": "Eruption gingivitis",
    "浆细胞龈炎": "Plasma cell gingivitis",
    "化脓性牙龈炎": "Suppurative gingivitis",
    "增生性牙龈炎": "Hyperplastic gingivitis",
    "溃疡性龈炎": "Ulcerative gingivitis",
    "边缘性龈炎": "Marginal gingivitis",
    "肥大性龈炎": "Hypertrophic gingivitis",
    "脱屑性龈炎": "Desquamative gingivitis",
    "急性多发性龈脓肿": "Acute multiple gingival abscesses",
    "牙周脓肿": "Periodontal abscess",
    "牙冠周脓肿": "Pericoronal abscess",
    "牙龈脓肿": "Gingival abscess",
    "急性冠周炎": "Acute pericoronitis",
    "复合性牙周炎": "Combined periodontitis",
    "单纯性牙周炎": "Simple periodontitis",
    "幼年牙周变性": "Juvenile periodontosis",
    "咬合创伤": "Occlusal trauma",
    "侵袭性牙周炎": "Aggressive periodontitis",
    "根分歧病变": "Furcation involvement",
    "种植体周围炎": "Peri-implantitis",
    "局部性牙龈退缩": "Localized gingival recession",
    "感染后牙龈退缩": "Postinfective gingival recession",
    "手术后牙龈退缩": "Postsurgical gingival recession",
    "牙龈增生": "Gingival hyperplasia",
    "药物性牙龈增生": "Drug-induced gingival hyperplasia",
    "遗传性龈纤维瘤病": "Hereditary gingival fibromatosis",
    "牙周巨细胞肉芽肿": "Peripheral giant cell granuloma",
    "牙龈黑斑": "Gingival melanotic macule",
    "牙龈粘膜色素沉着": "Gingival mucosal pigmentation",
    "种植体周围黏膜炎": "Peri-implant mucositis",
    "白血病的龈病损": "Leukemic gingival lesion",
    "上颌骨纤维增生": "Fibrous hyperplasia of maxilla",
    "下颌骨增生": "Mandibular hyperplasia",
    "巨上颌": "Macromaxilla",
    "小下颌": "Micrognathia of mandible",
    "小上颌": "Micrognathia of maxilla",
    "颏后缩": "Retrogenia",
    "方颏畸形": "Square chin deformity",
    "颏部畸形": "Chin deformity",
    "巨颌症": "Macrognathia",
    "颌骨发育不全": "Hypoplasia of jaw",
    "上颌骨骨质增生": "Hyperostosis of maxilla",
    "上颌骨发育不全": "Maxillary hypoplasia",
    "下颌骨骨质增生": "Hyperostosis of mandible",
    "下颌发育不全": "Mandibular hypoplasia",
    "下颌角肥大": "Hypertrophy of mandibular angle",
    "下颌角肥大伴咬肌肥大": "Hypertrophy of mandibular angle with masseter hypertrophy",
    "小颌畸形": "Micrognathia",
    "小颏畸形": "Microgenia",
    "唇腭裂术后颌骨发育不全": "Jaw hypoplasia after cleft lip and palate surgery",
    "错𬌗畸形骨性I类": "Skeletal Class I malocclusion",
    "错𬌗畸形骨性Ⅱ类": "Skeletal Class II malocclusion",
    "错𬌗畸形骨性Ⅲ类": "Skeletal Class III malocclusion",
    "颏部前突": "Progenia",
    "上颌前突下颌后缩": "Maxillary protrusion with mandibular retrusion",
    "上颌后缩下颌前突": "Maxillary retrusion with mandibular protrusion",
    "长面综合征": "Long face syndrome",
    "短面综合征": "Short face syndrome",
    "下颌前突偏斜": "Deviated mandibular prognathism",
    "偏颌畸形": "Jaw asymmetry",
    "双突颌畸形": "Bimaxillary protrusion",
    "上颌后缩": "Maxillary retrusion",
    "上颌前突": "Maxillary protrusion",
    "上下颌前突畸形": "Bimaxillary protrusion",
    "下颌后缩": "Mandibular retrusion",
    "下颌偏斜": "Mandibular deviation",
    "下颌前突": "Mandibular protrusion",
    "颌后缩": "Jaw retrusion",
    "颌骨不对称": "Jaw asymmetry",
    "后牙开𬌗": "Posterior open bite",
    "后牙锁合": "Posterior scissor bite",
    "前牙反𬌗": "Anterior crossbite",
    "深覆𬌗": "Deep overbite",
    "错𬌗畸形安氏I类": "Angle Class I malocclusion",
    "错𬌗畸形安氏Ⅱ类": "Angle Class II malocclusion",
    "错𬌗畸形安氏Ⅲ类": "Angle Class III malocclusion",
    "深覆盖": "Increased overjet",
    "覆咬合": "Overbite",
    "前牙开𬌗": "Anterior open bite",
    "牙弓中线偏离": "Dental midline deviation",
    "咬合异常": "Occlusal abnormality",
    "反𬌗": "Crossbite",
    "牙齿位置异常": "Abnormal tooth position",
    "牙齿间隙": "Spacing of teeth",
    "牙齿扭转": "Tooth rotation",
    "牙齿移位": "Tooth displacement",
    "第一恒磨牙异位萌出": "Ectopic eruption of first permanent molar",
    "牙的病理性移位": "Pathologic tooth migration",
    "牙错位": "Malposition of tooth",
    "牙列不齐": "Irregular dentition",
    "牙体缺损": "Tooth defect",
    "牙拥挤": "Dental crowding",
    "异位牙": "Ectopic tooth",
    "错𬌗畸形": "Malocclusion",
    "颌骨闭合异常": "Abnormal jaw closure",
    "颞下颌关节紊乱病": "Temporomandibular joint disorder",
    "颞下颌关节强直": "Temporomandibular joint ankylosis",
    "陈旧性颞下颌关节脱位": "Old temporomandibular joint dislocation",
    "颞颌关节综合征": "Temporomandibular joint syndrome",
    "颞颌关节骨关节病": "Temporomandibular joint osteoarthrosis",
    "颞颌关节炎": "Temporomandibular joint arthritis",
    "颜面部缺损": "Facial defect",
    "下颌畸形": "Mandibular deformity",
    "颌骨畸形": "Jaw deformity",
    "颌骨先天畸形": "Congenital deformity of jaw",
    "残留牙根": "Retained dental root",
    "单颌牙列缺失": "Edentulism of one jaw",
    "外伤性牙齿缺失": "Traumatic tooth loss",
    "后天性牙齿缺失": "Acquired tooth loss",
    "牙列部分缺失": "Partial edentulism",
    "无牙牙槽突萎缩": "Atrophy of edentulous alveolar process",
    "牙槽骨萎缩": "Alveolar bone atrophy",
    "牙槽嵴萎缩": "Alveolar ridge atrophy",
    "牙槽突萎缩": "Alveolar process atrophy",
    "残冠": "Residual crown",
    "牙痛": "Toothache",
    "牙槽嵴裂": "Alveolar cleft",
    "牙槽突不齐": "Irregular alveolar process",
    "牙槽嵴黏膜角化过度": "Hyperkeratosis of alveolar ridge mucosa",
    "牙槽出血": "Alveolar hemorrhage",
    "牙槽隐性裂": "Submucous alveolar cleft",
    "牙槽嵴增大": "Enlargement of alveolar ridge",
    "牙槽骨缺损": "Alveolar bone defect",
    "颌骨发育性牙源性囊肿": "Developmental odontogenic cyst of jaw",
    "颌骨始基囊肿": "Primordial cyst of jaw",
    "含牙囊肿": "Dentigerous cyst",
    "萌牙囊肿": "Eruption cyst",
    "牙龈囊肿": "Gingival cyst",
    "颌骨含牙囊肿": "Dentigerous cyst of jaw",
    "始基囊肿": "Primordial cyst",
    "腭骨囊肿": "Palatal cyst",
    "球上颌囊肿": "Globulomaxillary cyst",
    "口腔发育性（非牙源性）囊肿": "Developmental nonodontogenic cyst of oral region",
    "鼻腭囊肿": "Nasopalatine cyst",
    "鼻牙槽囊肿": "Nasoalveolar cyst",
    "鼻腭管囊肿": "Nasopalatine duct cyst",
    "颌骨囊肿": "Cyst of jaw",
    "颌出血性囊肿": "Hemorrhagic cyst of jaw",
    "颌动脉瘤性囊肿": "Aneurysmal cyst of jaw",
    "髁状突囊肿": "Condylar cyst",
    "上颌骨囊肿": "Maxillary cyst",
    "下颌骨囊肿": "Mandibular cyst",
    "爱泼斯坦小结[口底皮样囊肿]": "Epstein pearl [dermoid cyst of floor of mouth]",
    "腮腺淋巴上皮囊肿": "Lymphoepithelial cyst of parotid gland",
    "颏部皮样囊肿": "Dermoid cyst of chin",
    "颊囊肿": "Buccal cyst",
    "口腔表皮样囊肿": "Epidermoid cyst of oral cavity",
    "口腔皮样囊肿": "Dermoid cyst of oral cavity",
    "口腔黏液腺囊肿": "Mucous gland cyst of oral cavity",
    "口腔淋巴上皮囊肿": "Lymphoepithelial cyst of oral cavity",
    "口腔囊肿": "Oral cyst",
    "腭隆凸": "Torus palatinus",
    "颌的潜伏性骨囊肿": "Latent bone cyst of jaw",
    "斯塔夫尼囊肿": "Stafne bone cyst",
    "下颌隆凸": "Torus mandibularis",
    "腭裂手术后畸形": "Deformity after cleft palate surgery",
    "颌骨中枢性巨细胞病变": "Central giant cell lesion of jaw",
    "颌下区肉芽肿": "Granuloma of submandibular region",
    "颌骨巨细胞修复性肉芽肿": "Giant cell reparative granuloma of jaw",
    "颌骨巨细胞肉芽肿": "Giant cell granuloma of jaw",
    "颌肉芽肿": "Granuloma of jaw",
    "新生儿颌骨骨髓炎": "Neonatal osteomyelitis of jaw",
    "放射性颌骨坏死": "Osteoradionecrosis of jaw",
    "颌骨骨髓炎": "Osteomyelitis of jaw",
    "颌骨放射性骨髓炎": "Radiation osteomyelitis of jaw",
    "颌骨炎性增生": "Inflammatory hyperplasia of jaw",
    "颌骨骨炎": "Osteitis of jaw",
    "颌骨死骨": "Sequestrum of jaw",
    "化脓性颌骨髓炎": "Suppurative osteomyelitis of jaw",
    "髁状突炎": "Condylar inflammation",
    "慢性下颌骨边缘性骨髓炎": "Chronic peripheral osteomyelitis of mandible",
    "慢性下颌骨中央性骨髓炎": "Chronic central osteomyelitis of mandible",
    "慢性颌骨炎": "Chronic osteitis of jaw",
    "下颌炎性窦道": "Inflammatory sinus tract of mandible",
    "下颌骨局限坏死": "Localized necrosis of mandible",
    "翼腭窝炎": "Inflammation of pterygopalatine fossa",
    "颌骨牙槽炎": "Alveolitis of jaw",
    "牙槽骨骨炎": "Osteitis of alveolar bone",
    "干槽症": "Dry socket",
    "腭血肿": "Palatal hematoma",
    "颌骨纤维异常增殖症": "Fibrous dysplasia of jaw",
    "颌骨骨质增生": "Hyperostosis of jaw",
    "家族性巨颌症": "Cherubism",
    "颌骨单侧髁突增生": "Unilateral condylar hyperplasia of jaw",
    "颌骨单侧髁突发育不全": "Unilateral condylar hypoplasia of jaw",
    "髁突肥大": "Condylar hypertrophy",
    "单侧髁状突肥大": "Unilateral condylar hypertrophy",
    "髁状突骨疣": "Condylar osteophyte",
    "后天性腭畸形": "Acquired deformity of palate",
    "颌骨纤维结构发育不良": "Fibrous dysplasia of jaw",
    "颌外生性骨疣": "Exostosis of jaw",
    "上腭穿孔": "Palatal perforation",
    "颌部瘤样纤维组织增生": "Tumor-like fibrous hyperplasia of jaw",
    "颌骨缺损": "Defect of jaw",
    "下颌下腺良性增生": "Benign hyperplasia of submandibular gland",
    "唾液腺肥大": "Hypertrophy of salivary gland",
    "腮腺肥大": "Parotid hypertrophy",
    "颌下腺肥大": "Submandibular gland hypertrophy",
    "下颌下腺炎": "Submandibular sialadenitis",
    "舌下腺炎": "Sublingual sialadenitis",
    "慢性唾液腺炎": "Chronic sialadenitis",
    "硬化性唾液腺炎": "Sclerosing sialadenitis",
    "急性腮腺炎": "Acute parotitis",
    "急性颌下腺炎": "Acute submandibular sialadenitis",
    "急性舌下腺炎": "Acute sublingual sialadenitis",
    "慢性腮腺炎": "Chronic parotitis",
    "慢性颌下腺炎": "Chronic submandibular sialadenitis",
    "慢性舌下腺炎": "Chronic sublingual sialadenitis",
    "腮腺炎性假瘤": "Inflammatory pseudotumor of parotid gland",
    "硬化性涎腺炎": "Sclerosing sialadenitis",
    "阻塞性颌下腺炎": "Obstructive submandibular sialadenitis",
    "阻塞性腮腺炎": "Obstructive parotitis",
    "化脓性腮腺炎": "Suppurative parotitis",
    "涎腺脓肿": "Abscess of salivary gland",
    "腮腺脓肿": "Parotid abscess",
    "颌下腺脓肿": "Submandibular gland abscess",
    "舌下腺脓肿": "Sublingual gland abscess",
    "唾液导管瘘": "Salivary duct fistula",
    "腮腺瘘": "Parotid fistula",
    "腮腺导管瘘": "Parotid duct fistula",
    "颌下腺瘘": "Submandibular gland fistula",
    "舌下管结石": "Sublingual duct calculus",
    "唾液腺导管结石": "Salivary gland duct calculus",
    "腮腺导管结石": "Parotid duct calculus",
    "颌下腺导管结石": "Submandibular duct calculus",
    "口腔黏膜粘液囊肿": "Mucocele of oral mucosa",
    "舌下腺粘液囊肿": "Mucocele of sublingual gland",
    "腮腺囊肿": "Parotid cyst",
    "腮腺涎液潴留": "Salivary retention of parotid gland",
    "舌下腺囊肿": "Sublingual gland cyst",
    "舌下囊肿": "Ranula",
    "颌下腺囊肿": "Submandibular gland cyst",
    "颌下腺黏液囊肿": "Mucocele of submandibular gland",
    "唾液分泌紊乱": "Disturbance of salivary secretion",
    "唾液分泌过少": "Hyposalivation",
    "流涎症": "Sialorrhea",
    "口干燥症": "Xerostomia",
    "腮腺唾液潴留": "Salivary retention of parotid gland",
    "坏死性唾液腺化生": "Necrotizing sialometaplasia",
    "唾液腺肉芽肿": "Granuloma of salivary gland",
    "腮腺结节病": "Sarcoidosis of parotid gland",
    "涎腺管狭窄": "Stenosis of salivary duct",
    "米库利奇病": "Mikulicz disease",
    "腮腺管扩张": "Dilatation of parotid duct",
    "腮腺肉芽肿": "Granuloma of parotid gland",
    "涎腺良性淋巴上皮损害": "Benign lymphoepithelial lesion of salivary gland",
    "涎腺管扩张": "Dilatation of salivary duct",
    "涎腺导管阻塞": "Obstruction of salivary duct",
    "唾液腺病": "Sialadenosis",
    "轻型阿弗他溃疡": "Minor aphthous ulcer",
    "复发性坏死性黏膜腺周炎": "Recurrent necrotizing mucous gland periadenitis",
    "口腔阿弗他溃疡": "Oral aphthous ulcer",
    "疱疹样口炎": "Herpetiform stomatitis",
    "过敏性口炎": "Allergic stomatitis",
    "尼古丁口炎": "Nicotine stomatitis",
    "药物性口炎": "Drug-induced stomatitis",
    "糜烂性口炎": "Erosive stomatitis",
    "创伤性口腔黏膜溃疡": "Traumatic ulcer of oral mucosa",
    "腭部溃疡": "Palatal ulcer",
    "腭溃疡穿孔": "Perforated palatal ulcer",
    "腭部炎性假瘤": "Inflammatory pseudotumor of palate",
    "变应性口炎": "Allergic stomatitis",
    "溃疡性口炎": "Ulcerative stomatitis",
    "义齿性口炎": "Denture stomatitis",
    "口腔黏膜溃疡": "Ulcer of oral mucosa",
    "口底炎性假瘤": "Inflammatory pseudotumor of floor of mouth",
    "口腔感染": "Oral infection",
    "口腔炎": "Stomatitis",
    "小疱性口炎": "Vesicular stomatitis",
    "颊溃疡": "Buccal ulcer",
    "口腔炎性肿块": "Inflammatory mass of oral cavity",
    "上腭炎性肿物": "Inflammatory mass of palate",
    "颏下间隙感染": "Submental space infection",
    "口底多间隙感染": "Multiple space infection of floor of mouth",
    "颌下感染": "Submandibular infection",
    "口腔内脓肿": "Intraoral abscess",
    "软腭脓肿": "Abscess of soft palate",
    "硬腭脓肿": "Abscess of hard palate",
    "口腔脓肿": "Oral abscess",
    "颌下间隙感染": "Submandibular space infection",
    "颊部脓肿": "Buccal abscess",
    "颊间隙感染": "Buccal space infection",
    "颊瘘": "Buccal fistula",
    "腭瘘": "Palatal fistula",
    "颌下瘘管": "Submandibular fistula",
    "口腔瘘管": "Oral fistula",
    "口腔皮肤瘘": "Orocutaneous fistula",
    "眶下间隙感染": "Infraorbital space infection",
    "颞下间隙感染": "Infratemporal space infection",
    "舌下间隙感染": "Sublingual space infection",
    "咬肌间隙感染": "Masseteric space infection",
    "口蜂窝织炎": "Cellulitis of mouth",
    "翼下颌间隙感染": "Pterygomandibular space infection",
    "牙源性面部皮肤瘘": "Odontogenic cutaneous fistula of face",
    "颌面间隙感染": "Maxillofacial space infection",
    "颊黏膜脓肿": "Abscess of buccal mucosa",
    "药物性黏膜炎（口腔）（口咽）": "Drug-induced oral and oropharyngeal mucositis",
    "放射性黏膜炎（口腔）（口咽）": "Radiation-induced oral and oropharyngeal mucositis",
    "病毒性黏膜炎（口腔）（口咽）": "Viral oral and oropharyngeal mucositis",
    "腭黏膜炎": "Palatal mucositis",
    "感染性口角炎": "Infective angular cheilitis",
    "剥脱性唇炎": "Exfoliative cheilitis",
    "腺性唇炎": "Cheilitis glandularis",
    "唇表皮化": "Epidermization of lip",
    "唇黏液囊肿": "Mucocele of lip",
    "烧伤后唇畸形": "Post-burn deformity of lip",
    "变应性接触性唇炎": "Allergic contact cheilitis",
    "唇脓肿": "Lip abscess",
    "唇结节病": "Sarcoidosis of lip",
    "口角炎": "Angular cheilitis",
    "口腔灶性上皮增生": "Focal epithelial hyperplasia of oral mucosa",
    "烟斑": "Smoker's patch",
    "白色角化症": "White keratosis",
    "口腔黏膜红斑": "Erythroplakia of oral mucosa",
    "口腔黏膜白色水肿": "Leukoedema of oral mucosa",
    "颊鳞状上皮增生": "Squamous epithelial hyperplasia of buccal mucosa",
    "腭黏膜上皮增生": "Epithelial hyperplasia of palatal mucosa",
    "口腔黏膜结节病": "Sarcoidosis of oral mucosa",
    "口腔黏膜浆细胞肉芽肿": "Plasma cell granuloma of oral mucosa",
    "口腔黏膜化脓性肉芽肿": "Pyogenic granuloma of oral mucosa",
    "腭部黏膜下纤维化": "Submucous fibrosis of palate",
    "口腔黏膜炎性增生": "Inflammatory hyperplasia of oral mucosa",
    "腭部瘢痕": "Palatal scar",
    "腭黏膜炎症": "Inflammation of palatal mucosa",
    "后天性颊沟畸形": "Acquired buccal sulcus deformity",
    "口腔毛息肉": "Oral hairy polyp",
    "颊部炎症": "Inflammation of buccal region",
    "慢性颊黏膜下炎症": "Chronic submucosal inflammation of buccal mucosa",
    "口腔内血管增生": "Intraoral vascular hyperplasia",
    "口腔黏膜出血": "Hemorrhage of oral mucosa",
    "后天性软腭畸形": "Acquired deformity of soft palate",
    "翼沟过长": "Elongated pterygomandibular fold",
    "腭垂囊肿": "Cyst of uvula",
    "软腭肥厚": "Hypertrophy of soft palate",
    "腭麻痹": "Paralysis of palate",
    "软腭麻痹": "Paralysis of soft palate",
    "软腭震颤": "Tremor of soft palate",
    "金属引起的口腔黏膜病变": "Metal-induced oral mucosal lesion",
    "颊部息肉": "Polyp of buccal region",
    "舌创伤性溃疡": "Traumatic ulcer of tongue",
    "舌部嗜酸性溃疡": "Eosinophilic ulcer of tongue",
    "地图舌": "Geographic tongue",
    "毛舌": "Hairy tongue",
    "舌苔": "Coated tongue",
    "叶状乳头肥大": "Hypertrophy of foliate papillae",
    "光面舌": "Smooth tongue",
    "裂纹舌": "Fissured tongue",
    "沟纹舌": "Fissured tongue",
    "舌痛症": "Glossodynia",
    "舌瘘管": "Fistula of tongue",
    "舌畸形": "Deformity of tongue",
    "舌牙痕": "Scalloped tongue",
    "舌息肉": "Polyp of tongue",
    "舌粘连": "Tongue adhesion",
    "强直舌": "Ankyloglossia",
    "舌的静脉曲张": "Varices of tongue",
    "舌缺损": "Defect of tongue",
}

LOCATION_TERMS = {
    "口腔黏膜": "oral mucosa",
    "颊黏膜": "buccal mucosa",
    "腭黏膜": "palatal mucosa",
    "牙龈": "gingiva",
    "龈": "gingiva",
    "牙槽嵴": "alveolar ridge",
    "牙槽突": "alveolar process",
    "牙槽骨": "alveolar bone",
    "颞下颌关节": "temporomandibular joint",
    "下颌下腺": "submandibular gland",
    "颌下腺": "submandibular gland",
    "舌下腺": "sublingual gland",
    "唾液腺": "salivary gland",
    "涎腺": "salivary gland",
    "腮腺": "parotid gland",
    "上颌骨": "maxilla",
    "下颌骨": "mandible",
    "颌骨": "jaw",
    "下颌": "mandible",
    "上颌": "maxilla",
    "软腭": "soft palate",
    "硬腭": "hard palate",
    "腭垂": "uvula",
    "悬雍垂": "uvula",
    "口底": "floor of mouth",
    "颊部": "buccal region",
    "口腔": "oral cavity",
    "牙齿": "tooth",
    "牙": "tooth",
    "舌": "tongue",
    "唇": "lip",
    "腭": "palate",
    "颏部": "chin",
    "颏": "chin",
    "颌": "jaw",
}

PREFIX_TERMS = {
    "急性": "acute",
    "慢性": "chronic",
    "复发性": "recurrent",
    "遗传性": "hereditary",
    "药物性": "drug-induced",
    "放射性": "radiation-induced",
    "病毒性": "viral",
    "创伤性": "traumatic",
    "感染性": "infective",
    "变应性": "allergic",
    "过敏性": "allergic",
    "化脓性": "suppurative",
    "炎性": "inflammatory",
    "良性": "benign",
    "后天性": "acquired",
    "先天": "congenital",
    "局部性": "localized",
    "增生性": "hyperplastic",
    "糜烂性": "erosive",
    "溃疡性": "ulcerative",
}

SUFFIX_PATTERNS = [
    ("脓肿", "abscess of {loc}"),
    ("囊肿", "cyst of {loc}"),
    ("黏液囊肿", "mucocele of {loc}"),
    ("瘘管", "fistula of {loc}"),
    ("瘘", "fistula of {loc}"),
    ("溃疡", "ulcer of {loc}"),
    ("息肉", "polyp of {loc}"),
    ("肉芽肿", "granuloma of {loc}"),
    ("纤维瘤病", "fibromatosis of {loc}"),
    ("增生", "hyperplasia of {loc}"),
    ("肥大", "hypertrophy of {loc}"),
    ("萎缩", "atrophy of {loc}"),
    ("缺损", "defect of {loc}"),
    ("畸形", "deformity of {loc}"),
    ("肿物", "mass of {loc}"),
    ("瘢痕", "scar of {loc}"),
    ("出血", "hemorrhage of {loc}"),
    ("疼", "pain of {loc}"),
    ("痛", "pain of {loc}"),
    ("炎", "inflammation of {loc}"),
]


def canonical_subcategory_code(row: dict[str, str]) -> str:
    return row["subcategory_code"] + "00"


def is_who_exact(row: dict[str, str]) -> bool:
    return row["diagnosis_code"] == canonical_subcategory_code(row)


def strip_prefix(text: str) -> tuple[list[str], str]:
    prefixes: list[str] = []
    changed = True
    while changed:
        changed = False
        for cn, en in PREFIX_TERMS.items():
            if text.startswith(cn):
                prefixes.append(en)
                text = text[len(cn) :]
                changed = True
                break
    return prefixes, text


def translate_location(text: str) -> str | None:
    if not text:
        return None
    if text in LOCATION_TERMS:
        return LOCATION_TERMS[text]
    translated = text
    for cn in sorted(LOCATION_TERMS, key=len, reverse=True):
        translated = translated.replace(cn, LOCATION_TERMS[cn])
    if re.search(r"[\u3400-\u9fff]", translated):
        return None
    return translated


def conservative_translate(text: str, fallback: str) -> tuple[str, str]:
    text = text.replace("（", "(").replace("）", ")")
    prefixes, core = strip_prefix(text)
    for suffix, template in SUFFIX_PATTERNS:
        if core.endswith(suffix):
            loc = translate_location(core[: -len(suffix)])
            if loc:
                phrase = template.format(loc=loc)
                return (" ".join(prefixes + [phrase])).strip(), "CLINICAL_TRANSLATION"
    return fallback, "WHO_PARENT_INHERITED"


def semantic_mapping(row: dict[str, str]) -> tuple[str, str, str]:
    if is_who_exact(row):
        return row["who_name_en"], "WHO_EXACT", "WHO_ICD10"

    exact = SEMANTIC_OVERRIDES.get(row["diagnosis_name_cn"])
    if exact:
        mapping_type = (
            "LOCAL_SUBTYPE_TRANSLATION"
            if row["is_subtype"] == "TRUE"
            else "CLINICAL_STANDARD"
        )
        return exact, mapping_type, "DENTAL_TERMINOLOGY"

    translated, source_type = conservative_translate(
        row["diagnosis_name_cn"], row["structural_name_en"]
    )
    if row["is_subtype"] == "TRUE":
        return translated, "LOCAL_SUBTYPE_TRANSLATION", "CLINICAL_TRANSLATION"
    return translated, source_type, (
        "CLINICAL_TRANSLATION"
        if source_type == "CLINICAL_TRANSLATION"
        else "WHO_ICD10"
    )


def read_input(input_path: Path) -> list[dict[str, str]]:
    with input_path.open(encoding="utf-8", newline="") as input_file:
        reader = csv.DictReader(input_file, delimiter="\t")
        if reader.fieldnames != INPUT_COLUMNS:
            raise ValueError(
                f"Unexpected input columns: {reader.fieldnames}. "
                f"Expected: {INPUT_COLUMNS}"
            )
        return [dict(row) for row in reader]


def add_semantic_layer(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    output_rows: list[dict[str, str]] = []
    for source in rows:
        row = dict(source)
        row["who_icd_code"] = row["subcategory_code"]
        row["who_name_en"] = WHO_TERMS[row["subcategory_code"]]
        row["structural_name_en"] = row["diagnosis_name_en"]
        (
            row["semantic_name_en"],
            row["english_mapping_type"],
            row["semantic_source"],
        ) = semantic_mapping(row)
        output_rows.append(row)
    return output_rows


def write_tsv(rows: list[dict[str, str]], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.DictWriter(
            output_file,
            fieldnames=OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "K00-K14 Master v3"
    worksheet.append(OUTPUT_COLUMNS)
    for row in rows:
        worksheet.append([row[column] for column in OUTPUT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for index, column in enumerate(OUTPUT_COLUMNS, start=1):
        width = 18
        if column.endswith("_name_cn") or column.endswith("_name_en"):
            width = 38
        if column in {"chapter", "section", "structural_name_en", "semantic_name_en"}:
            width = 42
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    table = Table(displayName="K00K14MasterV3Semantic", ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    workbook.save(output_path)


def validate(input_rows: list[dict[str, str]], output_rows: list[dict[str, str]]) -> None:
    input_codes = [row["diagnosis_code"] for row in input_rows]
    output_codes = [row["diagnosis_code"] for row in output_rows]
    duplicate_codes = [
        code for code, count in Counter(output_codes).items() if count > 1
    ]
    missing_codes = [row for row in output_rows if not row["diagnosis_code"]]
    outside_range = [code for code in output_codes if not K00_K14_RE.match(code)]
    input_subtypes = {
        row["diagnosis_code"] for row in input_rows if row["is_subtype"] == "TRUE"
    }
    output_subtypes = {
        row["diagnosis_code"] for row in output_rows if row["is_subtype"] == "TRUE"
    }
    missing_semantic = [row for row in output_rows if not row["semantic_name_en"]]

    print(f"Input row count: {len(input_rows)}")
    print(f"Output row count: {len(output_rows)}")
    print(f"Same diagnosis_code sequence: {input_codes == output_codes}")
    print(f"Missing diagnosis_code: {len(missing_codes)}")
    print(f"Duplicate diagnosis_code: {len(duplicate_codes)}")
    print(f"Rows outside K00-K14: {len(outside_range)}")
    print(f"Dropped subtype rows: {len(input_subtypes - output_subtypes)}")
    print(f"Missing semantic_name_en: {len(missing_semantic)}")

    print("Counts by english_mapping_type:")
    for key, count in sorted(Counter(row["english_mapping_type"] for row in output_rows).items()):
        print(f"{key}\t{count}")
    print("Counts by english_mapping_confidence:")
    for key, count in sorted(
        Counter(row["english_mapping_confidence"] for row in output_rows).items()
    ):
        print(f"{key}\t{count}")
    print("Counts by semantic_source:")
    for key, count in sorted(Counter(row["semantic_source"] for row in output_rows).items()):
        print(f"{key}\t{count}")
    print("Counts by chapter_code:")
    for key, count in sorted(Counter(row["chapter_code"] for row in output_rows).items()):
        print(f"{key}\t{count}")


def validate_xlsx(output_path: Path) -> None:
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    worksheet = workbook["K00-K14 Master v3"]
    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    print(f"XLSX columns match: {header == OUTPUT_COLUMNS}")
    print(f"XLSX row count: {len(rows)}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build K00-K14 v3 semantic English refinement layer."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--tsv", type=Path, default=DEFAULT_TSV)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()

    input_rows = read_input(args.input)
    output_rows = add_semantic_layer(input_rows)
    write_tsv(output_rows, args.tsv)
    write_xlsx(output_rows, args.xlsx)
    print(f"Wrote {args.tsv}")
    print(f"Wrote {args.xlsx}")
    validate(input_rows, output_rows)
    validate_xlsx(args.xlsx)


if __name__ == "__main__":
    main()
