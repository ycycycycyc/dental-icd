#!/usr/bin/env python3
"""Create English snake_case column versions of the K00-K14 master table."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


INPUT_COLUMNS = [
    "章",
    "节",
    "类目代码",
    "类目名称",
    "亚目代码",
    "亚目名称",
    "诊断代码",
    "诊断名称",
    "chapter_code",
    "parent_code",
    "is_subtype",
    "subtype_number",
    "code_level",
]

OUTPUT_COLUMNS = [
    "chapter",
    "section",
    "category_code",
    "category_name_cn",
    "subcategory_code",
    "subcategory_name_cn",
    "diagnosis_code",
    "diagnosis_name_cn",
    "chapter_code",
    "parent_code",
    "is_subtype",
    "subtype_number",
    "code_level",
]

DEFAULT_INPUT = Path("K00-K14_master_table_v1.tsv")
DEFAULT_TSV = Path("K00-K14_master_table_v1_english_columns.tsv")
DEFAULT_XLSX = Path("K00-K14_master_table_v1_english_columns.xlsx")


def read_rows(input_path: Path) -> list[dict[str, str]]:
    with input_path.open(encoding="utf-8", newline="") as input_file:
        reader = csv.DictReader(input_file, delimiter="\t")
        if reader.fieldnames != INPUT_COLUMNS:
            raise ValueError(
                f"Unexpected input columns: {reader.fieldnames}. "
                f"Expected: {INPUT_COLUMNS}"
            )
        return [
            {
                output_column: row[input_column]
                for input_column, output_column in zip(INPUT_COLUMNS, OUTPUT_COLUMNS)
            }
            for row in reader
        ]


def write_tsv(rows: list[dict[str, str]], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.DictWriter(
            output_file,
            fieldnames=OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "K00-K14 Master"
    worksheet.append(OUTPUT_COLUMNS)
    for row in rows:
        worksheet.append([row[column] for column in OUTPUT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 28,
        "B": 28,
        "C": 16,
        "D": 24,
        "E": 18,
        "F": 34,
        "G": 18,
        "H": 38,
        "I": 14,
        "J": 14,
        "K": 12,
        "L": 15,
        "M": 14,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    table = Table(displayName="K00K14MasterEnglishColumns", ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    workbook.save(output_path)


def validate(tsv_path: Path, xlsx_path: Path) -> None:
    with tsv_path.open(encoding="utf-8", newline="") as input_file:
        rows = list(csv.DictReader(input_file, delimiter="\t"))
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook["K00-K14 Master"]
    xlsx_header = [
        cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
    ]
    xlsx_rows = list(worksheet.iter_rows(min_row=2, values_only=True))

    print(f"TSV columns match: {list(rows[0].keys()) == OUTPUT_COLUMNS}")
    print(f"TSV row count: {len(rows)}")
    print(f"XLSX columns match: {xlsx_header == OUTPUT_COLUMNS}")
    print(f"XLSX row count: {len(xlsx_rows)}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Rename K00-K14 master table columns to English snake_case."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--tsv", type=Path, default=DEFAULT_TSV)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()

    rows = read_rows(args.input)
    write_tsv(rows, args.tsv)
    write_xlsx(rows, args.xlsx)
    print(f"Wrote {args.tsv}")
    print(f"Wrote {args.xlsx}")
    validate(args.tsv, args.xlsx)


if __name__ == "__main__":
    main()
